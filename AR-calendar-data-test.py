import pandas as pd
from datetime import datetime, timedelta
import sqlite3
import os
import sys

import openpyxl
from openpyxl.styles import PatternFill, Border, Side, numbers
import re

import streamlit as st
from PyInstaller.utils.hooks import copy_metadata

datas = copy_metadata("streamlit")

st.title("AR Calendar Auto Complete")
years = list(range(2000, 2081))
months = list(range(1, 13))

default_start_month_index = months.index(12) # If a new excel document is used please change the start date here.
default_start_year_index = years.index(2023)

selected_start_month = st.selectbox("Select Month at the start of excel file", months, index=default_start_month_index)
selected_start_year = st.selectbox("Select Year at the start of excel file", years, index=default_start_year_index)

today = datetime.now().date()
first = today.replace(day=1)
monthToChange = first - timedelta(days=1)

selected_month = st.selectbox("Select Month to analyse", months, index=months.index(monthToChange.month))
selected_year = st.selectbox("Select Year to analyse", years, index=years.index(monthToChange.year))

monthToChange = monthToChange.replace(month=selected_month, year=selected_year)

# dateFormat = r'\d{2}/\d{2}/\d{4}'
# startDateExcelFile = datetime.strptime("01/12/2023", '%d/%m/%Y')
if selected_start_month < 10:
    startDateExcelFile = datetime.strptime(f"01/0{selected_start_month}/{selected_start_year}", '%d/%m/%Y')
else:
    startDateExcelFile = datetime.strptime(f"01/{selected_start_month}/{selected_start_year}", '%d/%m/%Y')

csvLocation = st.text_input("Enter csv files address for AR registed users (remember the .csv)")
excelFileLocation = st.text_input("Enter Excel files address for the AR calender (remember the .xlsx)")

submitted = st.button("Submit")

def do_analysis():
    #Read the csv file into a pandas dataframe
    df = pd.read_csv(csvLocation, skiprows=1, encoding='ISO-8859-1', index_col=False)
    df.columns = df.columns.str.replace(" ", "_")
    df.replace('-', '', inplace=True)

    slectedColumns = ['Investor_Id', 'Create_Date', 'Date_Last_Took_App_Test', 'Number_App_Test_Fails', 'Number_App_Test_Passes', 'Last_Login', 'Last_Investment', 
        'Categorisation', 'Email_Address', 'Email_Confirmed', 'Nationality', 'Resident_Country', 'Test_Investor', 'User_Kyc_Status', 'Pep', 'Vulnerable_Customer']
    monitoring_df=df[slectedColumns].copy()

    # print(monitoring_df)

    #monitoring_df.head()

    #View first 5 rows
    # print(monitoring_df.dtypes)
    monitoring_df['Investor_Id'] = monitoring_df['Investor_Id'].astype(int)
    monitoring_df['Create_Date'] = pd.to_datetime(monitoring_df['Create_Date'],format="%d/%m/%Y")
    monitoring_df['Date_Last_Took_App_Test'] = pd.to_datetime(monitoring_df['Date_Last_Took_App_Test'], format="%d/%m/%Y")
    monitoring_df['Last_Login'] = pd.to_datetime(monitoring_df['Last_Login'], format="%d/%m/%Y")
    monitoring_df['Last_Investment'] = pd.to_datetime(monitoring_df['Last_Investment'], format="%d/%m/%Y")
    monitoring_df['Number_App_Test_Fails'] = monitoring_df['Number_App_Test_Fails'].astype(int)
    monitoring_df['Number_App_Test_Passes'] = monitoring_df['Number_App_Test_Passes'].astype(int)
    # print(monitoring_df.dtypes)
    database = 'database.db'

    if os.path.exists(database):
        os.remove(database)
    # Creating SQLite database
    conn = sqlite3.connect(database)
    cursor = conn.cursor()

    # # Create SQLite table - Table Definition
    create_table = '''CREATE TABLE IF NOT EXISTS ARMonitoringData(
                    Investor_Id INTEGER PRIMARY KEY NOT NULL,
                    Create_Date datetime NOT NULL,
                    Date_Last_Took_App_Test datetime,
                    Last_Login datetime,
                    Last_Investment datetime,
                    Number_App_Test_Fails INTEGER,
                    Number_App_Test_Passes INTEGER,
                    Categorisation VARCHAR(40),
                    Email_Address VARCHAR(320),
                    Email_Confirmed VARCHAR(5),
                    Nationality VARCHAR(40),
                    Resident_Country VARCHAR(40),
                    Test_Investor VARCHAR(5),
                    User_Kyc_Status VARCHAR(40),
                    Pep VARCHAR(5),
                    Vulnerable_Customer VARCHAR(5));
                    '''

    # Creating the table into our database
    cursor.execute(create_table)

    # insert the data from the pandas DataFrame into the SQLite table
    monitoring_df.to_sql('ARMonitoringData', conn, if_exists='replace', index = False)

    # print(pd.read_sql("""SELECT COUNT(*) 
    #     FROM ARMonitoringData 
    #     WHERE Email_Address LIKE '%test%' OR Email_Address LIKE '%%sharein%';""", conn).iloc[0, 0])

    def removeTesters():
        deletion_query = """DELETE FROM ARMonitoringData 
        WHERE Email_Address LIKE '%test%' OR Email_Address LIKE '%%sharein%';"""

        conn.execute(deletion_query)

    removeTesters()

    workbook = openpyxl.load_workbook(excelFileLocation)
    sheet = workbook.active
    border = Border(left=Side(style='thin', color='000000'), 
                    right=Side(style='thin', color='000000'), 
                    top=Side(style='thin', color='000000'), 
                    bottom=Side(style='thin', color='000000'))

    if (monthToChange.year*12 + monthToChange.month >= startDateExcelFile.year*12 + startDateExcelFile.month):
        COLUMN_OFFSET = 3
        ASCII_OF_A = 65

        letterForMonthInt = (monthToChange.year - startDateExcelFile.year)*12 + (monthToChange.month - startDateExcelFile.month)
        letterForMonthInt += ASCII_OF_A + COLUMN_OFFSET
        if letterForMonthInt < ASCII_OF_A + 26: # If the column is one character long.
            cellColumnForMonth = chr(letterForMonthInt)
        elif letterForMonthInt < ASCII_OF_A + 26*26: # If the column is two characters long.
            cellColumnForMonth = str(chr(int((letterForMonthInt-ASCII_OF_A)/26) + ASCII_OF_A - 1)) + str(chr(int((letterForMonthInt-ASCII_OF_A)%26) + ASCII_OF_A))
        else:
            print("Error: The month entered is two far into the future. Date entered is more than 56 years ahead of the start date")
    else:
        print(f"Error: The inputted date is before {startDateExcelFile}.")
        sys.exit(1)

    def paintCell(cell, colour):
        if colour == "green":
            cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        elif colour == "yellow":
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        elif colour == "amber":
            cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        elif colour == "red":
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        elif colour == "dark red":
            cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        elif colour == "white":
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        else:
            print("Error: Invalid colour. The colour must be one of green, yellow, amber, red, dark red, and white.")
            sys.exit(1)
        cell.border = border

    def writeToCellPercent(cell, value):
        cell.value = value
        cell.number_format = numbers.FORMAT_PERCENTAGE

    def paintDependingOnPrevious(sheet, cell, column, row, increase): 
        prevCellPos = column[:-1] + chr(ord(column[-1]) - 1) + row # So that it works when the column is more than one char
        previousMonthVal = sheet[prevCellPos].value
        monthVal = round(sheet[column + row].value, 2)

        if previousMonthVal != None:
            if increase:
                if monthVal - previousMonthVal >= 0.03:
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
            else:
                if previousMonthVal - monthVal >= 0.03:
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
            cell.border = border

    def checkIfZero(num):
        if num == 0:
            print("Error: You are trying to divide by zero when calculating one of the percentages. This is because no people have registered within a time frame. " +
            "The month entered has not happened yet or the csv is empty (probaly one of these reasons).")
            sys.exit(1)

    # Printing pandas dataframe
    # print(pd.read_sql('''SELECT * FROM ARMonitoringData limit 2''', conn))

    #AR calendar check 1.2	"What percentage of the ARs investor base represent overseas residents?

    # df_item1_2 = pd.read_sql('''SELECT distinct Investor_Id, Resident_Country FROM ARMonitoringData''', conn)
    # print(df_item1_2)

    # print(df[df["Test_Investor"] != "No"]["Test_Investor"])
    # print(df[(df["Categorisation"] != "SelfCertifiedSophisticatedInvestor") & (df["Categorisation"] != "RestrictedInvestor") & 
    # (df["Categorisation"] != "CertifiedHNWInvestor") & (df["Categorisation"] != "")]["Categorisation"])

    # print(df[((df["User_Kyc_Status"] != "Regular") & (df["User_Kyc_Status"] != "Light"))]["User_Kyc_Status"])

    # unique_investors = df.Investor_Id.unique()
    # num_investors = len(unique_investors)

    num_overseas = pd.read_sql(
        """SELECT COUNT(*)
        FROM ARMonitoringData
        WHERE Resident_Country != 'United Kingdom' AND User_Kyc_Status = "Regular";""", conn).iloc[0, 0]

    num_investors = pd.read_sql(
        """SELECT COUNT(*)
        FROM ARMonitoringData
        WHERE User_Kyc_Status = "Regular";""", conn).iloc[0, 0]

    checkIfZero(num_investors)

    percent_overseas_1_2 = num_overseas/num_investors
    # print(percent_overseas_1_2)


    if (percent_overseas_1_2 < 0):
        print("Error: The percentage of overseas residents cannot be negative")
        sys.exit(1)
    else:
        cellPos = cellColumnForMonth + "4"
        cell = sheet[cellPos]
        if (percent_overseas_1_2 <= 0.1):
            paintCell(cell, "green")
        elif (percent_overseas_1_2 <= 0.2):
            paintCell(cell, "yellow")
        elif (percent_overseas_1_2 <= 0.3):
            paintCell(cell, "amber")
        else:
            paintCell(cell, "red")
            # print("Error: The percentage of overseas residents has exceeded 30%%, colour is unspecified")
            # sys.exit(1)
        writeToCellPercent(cell, percent_overseas_1_2)


    # """
    # 0-10% - green
    # 11-20% - yellow
    # 21-30% - amber"""

    #1.3	"What percentage of the ARs investor base represent residents of FATF grey-list countries?
    # """
    # 0-10% of investor base - green
    # 11-15% of investor base - yellow
    # 16-20% of investor base - amber"""

    num_grey_list = pd.read_sql(
        """SELECT COUNT(*)
        FROM ARMonitoringData
        WHERE Resident_Country IN ('Bulgaria', 'Burkina Faso', 'Cameroon', 'Croatia', 'Congo, the Democratic Republic of the', 
        'Haiti', 'Jamaica', 'Kenya', 'Mali', 'Mozambique', 'Namibia', 'Nigeria', 'Philippines', 'Senegal', 'South Africa', 
        'South Sudan', 'Syrian Arab Republic', 'Tanzania, United Republic of', 'Turkey', 'Vietnam', 'Yemen') AND User_Kyc_Status = "Regular";""", conn).iloc[0, 0]

    percent_grey_list_1_3 = num_grey_list / num_investors

    if (percent_grey_list_1_3 < 0):
        print("Error: The percentage of grey-list country residents cannot be negative")
        sys.exit(1)
    else:
        cellPos = cellColumnForMonth + "5"
        cell = sheet[cellPos]
        if (percent_grey_list_1_3 <= 0.1):
            paintCell(cell, "green")
        elif (percent_grey_list_1_3 <= 0.15):
            paintCell(cell, "yellow")
        elif (percent_grey_list_1_3 <= 0.2):
            paintCell(cell, "amber")
        else:
            paintCell(cell, "red")
            # print("Error: The percentage of grey-list country residents has exceeded 20%%, colour is unspecified")
            # sys.exit(1)
        writeToCellPercent(cell, percent_grey_list_1_3)



    #1.4	"What percentage of the ARs investor base have been identifed as a PEP / RCA of a PEP?
    # """
    # 0% of investor base - green
    # >0% - 1% of investor base - yellow
    # 2-5% of investor base - amber
    # 6-9% of investor base - red
    # 10%+ of investor base - dark red"""

    # peps = pd.read_sql(
    #     """SELECT Investor_Id, Pep
    #     FROM ARMonitoringData;""", conn)
    # print(peps)

    num_pep = pd.read_sql( # I am assuming that the field Pep identifies whether they are a PEP / RCA of a PEP
        """SELECT COUNT(*)
        FROM ARMonitoringData
        WHERE Pep = 'Yes';""", conn).iloc[0, 0]

    percent_PEP_or_RCA_1_4 = num_pep / num_investors

    if (percent_PEP_or_RCA_1_4 < 0):
        print("Error: The percentage of investors that have been identifed as a PEP / RCA of a PEP cannot be negative")
        sys.exit(1)
    else:
        cellPos = cellColumnForMonth + "6"
        cell = sheet[cellPos]
        if (percent_PEP_or_RCA_1_4 == 0):
            paintCell(cell, "green")
        elif (percent_PEP_or_RCA_1_4 <= 0.01):
            paintCell(cell, "yellow")
        elif (percent_PEP_or_RCA_1_4 <= 0.05):
            paintCell(cell, "amber")
        elif (percent_PEP_or_RCA_1_4 <= 0.09):
            paintCell(cell, "red")
        else:
            paintCell(cell, "dark red")
        writeToCellPercent(cell, percent_PEP_or_RCA_1_4)

    #1.5	"What percentage of the ARs investor base have been identifed as vulnerable?
    # """
    # 0% of investor base - green
    # >0% - 1% of investor base - yellow
    # 2-5% of investor base - amber
    # 6-9% of investor base - red
    # 10%+ of investor base - dark red"""

    # print(df["Vulnerable_Customer"][:5])
    num_vulnerable = pd.read_sql(
        """SELECT COUNT(*)
        FROM ARMonitoringData
        WHERE Vulnerable_Customer = 'Yes';""", conn).iloc[0, 0]

    percent_vulnerable_1_5 = num_vulnerable / num_investors

    if (percent_vulnerable_1_5 < 0):
        print("Error: The percentage of investors that have been identifed as vulnerable cannot be negative")
        sys.exit(1)
    else:
        cellPos = cellColumnForMonth + "8"
        cell = sheet[cellPos]
        if (percent_vulnerable_1_5 == 0):
            paintCell(cell, "green")
        elif (percent_vulnerable_1_5 <= 0.01):
            paintCell(cell, "yellow")
        elif (percent_vulnerable_1_5 <= 0.05):
            paintCell(cell, "amber")
        elif (percent_vulnerable_1_5 <= 0.09):
            paintCell(cell, "red")
        else:
            paintCell(cell, "dark red")
        writeToCellPercent(cell, percent_vulnerable_1_5)

    #12.1	"How many new users attempted but never passed the app test in the past month? What is this as percentage of new registered users on the platform? 
    #Aim: to identify ARs whose app tests are not sufficiently effective at screening out individuals for whom the financial products may not be suitable. 
    # """
    # <15% - Green
    # >15% - Yellow"""
    def convert_to_yyyy_mm(date):
        month = date.month
        strMonth = str(month)
        if month <= 9:
            strMonth = "0" + strMonth
        
        return str(date.year) + "-" + strMonth

    num_not_passed_app_test = pd.read_sql( 
        """SELECT COUNT(*)
        FROM ARMonitoringData
        WHERE strftime('%Y-%m', Date_Last_Took_App_Test) = '""" + convert_to_yyyy_mm(monthToChange) + "' AND Number_App_Test_Passes = 0;", conn).iloc[0, 0]

    # Make sure they have taken the app test using Date Last Took App Test

    # print(monitoring_df["Number_App_Test_Passes"])

    num_registered = pd.read_sql( 
        """SELECT COUNT(*)
        FROM ARMonitoringData
        WHERE strftime('%Y-%m', Date_Last_Took_App_Test) = '""" + convert_to_yyyy_mm(monthToChange) + "';", conn).iloc[0, 0]

    # num_registered = pd.read_sql( 
    #     """SELECT COUNT(*)
    #     FROM ARMonitoringData
    #     WHERE strftime('%Y-%m', Create_Date) = '""" + convert_to_yyyy_mm(monthToChange) + "';", conn).iloc[0, 0]

    # add email confirmed

    checkIfZero(num_registered)
    # print(num_not_passed_app_test)
    # print("\n")
    # print(num_registered)


    percent_not_passed_12_1 = num_not_passed_app_test / num_registered

    if (percent_not_passed_12_1 < 0):
        print("Error: The percentage of investors in the last month that have attempted but never passed the app test cannot be negative")
        sys.exit(1)
    else:
        cellPos = cellColumnForMonth + "58"
        cell = sheet[cellPos]
        if (percent_not_passed_12_1 <= 0.15):
            paintCell(cell, "green")
        else:
            paintCell(cell, "yellow")
        writeToCellPercent(cell, percent_not_passed_12_1)

    #12.2	"How many new users passed the app test on the first attempt in the last month? What is this as percentage of new registered users on the platform?
    #Aim: to understand the proportion of new investors on the platform who appear to have a ‘confident’ grasp of the risks and products.
    # """
    # >65%+ - Green
    # < 65% - Yellow"""

    num_passed_first_try = pd.read_sql( 
        """SELECT COUNT(*)
        FROM ARMonitoringData
        WHERE strftime('%Y-%m', Date_Last_Took_App_Test) = '""" + convert_to_yyyy_mm(monthToChange) + "' AND Number_App_Test_Passes != 0 AND Number_App_Test_Fails = 0;", 
        conn).iloc[0, 0]

    # print(num_passed_first_try)

    percent_passed_first_try_12_2 = num_passed_first_try / num_registered

    if (percent_passed_first_try_12_2 < 0):
        print("Error: The percentage of investors in the last month that passed the app test on there first try cannot be negative")
        sys.exit(1)
    else:
        cellPos = cellColumnForMonth + "59"
        cell = sheet[cellPos]
        if (percent_passed_first_try_12_2 <= 0.65):
            paintCell(cell, "yellow")
        else:
            paintCell(cell, "green")
        writeToCellPercent(cell, percent_passed_first_try_12_2)

    #12.3	"How many new users who registered on the platform in the last month passed the app test after previously failing it? What is this as percentage of new registered users on the platform?
    #Aim: to identify potential risk to ShareIn of baseline vs possible increase in the investor base who may be ‘borderline’ in their understanding of the products and risks and who may therefore need more support or be more likely to complain in future. 
    # """
    # <15%  - Green
    # >15% - Yellow"""

    num_passed_not_first_try = pd.read_sql( 
        """SELECT COUNT(*)
        FROM ARMonitoringData
        WHERE strftime('%Y-%m', Date_Last_Took_App_Test) = '""" + convert_to_yyyy_mm(monthToChange) + "' AND Number_App_Test_Passes != 0 AND Number_App_Test_Fails != 0;", 
        conn).iloc[0, 0]

    percent_passed_not_first_try_12_3 = num_passed_not_first_try / num_registered

    if (percent_passed_not_first_try_12_3 < 0):
        print("Error: The percentage of investors in the last month that passed the app test not on there first try cannot be negative")
        sys.exit(1)
    else:
        cellPos = cellColumnForMonth + "60"
        cell = sheet[cellPos]
        if (percent_passed_not_first_try_12_3 <= 0.15):
            paintCell(cell, "green")
        else:
            paintCell(cell, "yellow")
        writeToCellPercent(cell, percent_passed_not_first_try_12_3)

    #12.4	What percentage of fully onboarded, restricted investors are in the AR database? 
    # print(df["User_Kyc_Status"][:100])

    num_restricted = pd.read_sql( 
        """SELECT COUNT(*)
        FROM ARMonitoringData
        WHERE User_Kyc_Status = 'Regular' AND (Categorisation = 'RestrictedInvestor' OR Categorisation = 'RestrictedCompany');""", 
        conn).iloc[0, 0]

    num_onboarded = pd.read_sql( 
        """SELECT COUNT(*)
        FROM ARMonitoringData
        WHERE User_Kyc_Status = 'Regular';""", 
        conn).iloc[0, 0]


    percent_restricted_12_4 = num_restricted / num_onboarded

    if (percent_restricted_12_4 < 0):
        print("Error: The percentage of investors that are restricted cannot be negative")
        sys.exit(1)
    else:
        cellPos = cellColumnForMonth + "61"
        cell = sheet[cellPos]
        paintCell(cell, "green")
        writeToCellPercent(cell, percent_restricted_12_4)
    #12.5	What percentage of fully onboarded, HNW investors are in the AR database? 

    num_HNW = pd.read_sql( 
        """SELECT COUNT(*)
        FROM ARMonitoringData
        WHERE User_Kyc_Status = 'Regular' AND (Categorisation = 'CertifiedHNWInvestor' OR Categorisation = 'HNWCompany');""", 
        conn).iloc[0, 0]

    percent_HNW_12_5 = num_HNW / num_onboarded

    if (percent_HNW_12_5 < 0):
        print("Error: The percentage of investors that are High Net Worth cannot be negative")
        sys.exit(1)
    else:
        cellPos = cellColumnForMonth + "62"
        cell = sheet[cellPos]
        paintCell(cell, "green")
        writeToCellPercent(cell, percent_HNW_12_5)
    #12.6	What percentage of fully onboarded, SCS investors are in the AR database? 
    num_SCS = pd.read_sql( 
        """SELECT COUNT(*)
        FROM ARMonitoringData
        WHERE User_Kyc_Status = 'Regular' AND (Categorisation = 'SelfCertifiedSophisticatedInvestor' OR Categorisation = 'SCSICompany');""", 
        conn).iloc[0, 0]

    percent_SCS_12_6 = num_SCS / num_onboarded

    if (percent_SCS_12_6 < 0):
        print("Error: The percentage of investors that are Self Certified Sophisticated Investors cannot be negative")
        sys.exit(1)
    else:
        cellPos = cellColumnForMonth + "63"
        cell = sheet[cellPos]
        paintCell(cell, "green")
        writeToCellPercent(cell, percent_SCS_12_6)
    #12.7	"Of those investors who have invested in the last month, what percentage are restricted?
    # print(df["Last_Investment"][:100])
    num_restricted_invested_in_month = pd.read_sql( 
        """SELECT COUNT(*)
        FROM ARMonitoringData
        WHERE strftime('%Y-%m', Last_Investment) = '""" + convert_to_yyyy_mm(monthToChange) + "' AND (Categorisation = 'RestrictedInvestor' OR Categorisation = 'RestrictedCompany');", 
        conn).iloc[0, 0]

    # Include companies

    num_invested_in_month = pd.read_sql( 
        """SELECT COUNT(*)
        FROM ARMonitoringData
        WHERE strftime('%Y-%m', Last_Investment) = '""" + convert_to_yyyy_mm(monthToChange) + "';", 
        conn).iloc[0, 0]

    percent_restricted_invested_12_7 = num_restricted_invested_in_month / num_invested_in_month

    if (percent_restricted_invested_12_7 < 0):
        print("Error: The percentage of investors that are restricted cannot be negative")
        sys.exit(1)
    else:
        row = "64"
        cellPos = cellColumnForMonth + row
        cell = sheet[cellPos]
        writeToCellPercent(cell, percent_restricted_invested_12_7)
        paintDependingOnPrevious(sheet, cell, cellColumnForMonth, row, False)
    # If the percentage change has decreased more than 3% from the previous month - Yellow"
    #12.8	"Of those investors who have invested in the last month, what percentage are HNW?

    num_HNW_invested_in_month = pd.read_sql( 
        """SELECT COUNT(*)
        FROM ARMonitoringData
        WHERE strftime('%Y-%m', Last_Investment) = '""" + convert_to_yyyy_mm(monthToChange) + "' AND (Categorisation = 'CertifiedHNWInvestor' OR Categorisation = 'HNWCompany');", 
        conn).iloc[0, 0]

    percent_HNW_invested_12_8 = num_HNW_invested_in_month / num_invested_in_month

    if (percent_HNW_invested_12_8 < 0):
        print("Error: The percentage of investors that are High Net Worth cannot be negative")
        sys.exit(1)
    else:
        row = "65"
        cellPos = cellColumnForMonth + row
        cell = sheet[cellPos]
        writeToCellPercent(cell, percent_HNW_invested_12_8)
        paintDependingOnPrevious(sheet, cell, cellColumnForMonth, row, True)
    # If the percentage change has increased more than 3% from the previous month - Yellow"
    #12.9	"Of those investors who have invested in the last month, what percentage are SCS?

    num_SCS_invested_in_month = pd.read_sql( 
        """SELECT COUNT(*)
        FROM ARMonitoringData
        WHERE strftime('%Y-%m', Last_Investment) = '""" + convert_to_yyyy_mm(monthToChange) + "' AND (Categorisation = 'SelfCertifiedSophisticatedInvestor'" +
        " OR Categorisation = 'SCSICompany');", conn).iloc[0, 0]

    percent_SCS_invested_12_9 = num_SCS_invested_in_month / num_invested_in_month

    if (percent_SCS_invested_12_9 < 0):
        print("Error: The percentage of investors that are Self Certified Sophisticated Investors cannot be negative")
        sys.exit(1)
    else:
        row = "66"
        cellPos = cellColumnForMonth + row
        cell = sheet[cellPos]
        writeToCellPercent(cell, percent_SCS_invested_12_9)
        paintDependingOnPrevious(sheet, cell, cellColumnForMonth, row, True)
    # If the percentage change has increased more than 3% from the previous month - Yellow"
    #12.10	"What percentage of the AR database are Trusts?
    # """
    # 0% of investor base - green
    # >0% of investor base - yellow
    # 2-5% of investor base - amber
    # 5-10% of investor base - red
    # 10+ of investor base - dark red
    # """

    num_trust = pd.read_sql(
        """SELECT COUNT(*)
        FROM ARMonitoringData
        WHERE Categorisation IN ('SCSITrust', 'AdvisedInvestorTrust', 'SCSITrustEquityAndP2P', 'HighValueTrust', 'SCSITrustP2POnly');""", conn).iloc[0, 0]

    percent_trust_12_10 = num_trust / num_investors

    if (percent_trust_12_10 < 0):
        print("Error: The percentage of investors are trusts cannot be negative")
        sys.exit(1)
    else:
        cellPos = cellColumnForMonth + "67"
        cell = sheet[cellPos]
        if (percent_trust_12_10 == 0):
            paintCell(cell, "green")
        elif (percent_trust_12_10 < 0.02):
            paintCell(cell, "yellow")
        elif (percent_trust_12_10 <= 0.05):
            paintCell(cell, "amber")
        elif (percent_trust_12_10 <= 0.1):
            paintCell(cell, "red")
        else:
            paintCell(cell, "dark red")
        writeToCellPercent(cell, percent_trust_12_10)
    #xslx writer https://pypi.org/project/XlsxWriter/
    # write output to Excel
    #Ideal - append to existing calendar in dropbox - hard to do. Could have files written locally and manually move to dropbox. Team can copy/paste
    conn.commit()
    conn.close()
    try:
        workbook.save(excelFileLocation)
    except PermissionError:
        print("Error: Premission denied accessing the Excel file. Please make sure that the excel file in question is closed before running the program.")

if submitted:
    do_analysis()